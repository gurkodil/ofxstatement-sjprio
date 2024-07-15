from typing import Iterable

import re
import itertools

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id
from openpyxl import load_workbook
from typing import Any

import logging


def take(n: int, iterable: Iterable[Any]):
    return list(itertools.islice(iterable, n))


class SJPrioPlugin(Plugin):
    """SJPrio Bank <https://www.sj.se/sj-prio>"""

    def get_parser(self, filename: str) -> "SJPrioParser":
        return SJPrioParser(filename)


class SJPrioParser(StatementParser[str]):
    ACCOUNT_REGEX = "^Kortnummer (\d{6}\*{6}\d{4})$"
    TRANSACTION_DATE = 0
    TRANSACTION_DESCRIPTION = 2
    TRANSACTION_LOCATION = 3
    TRANSACTION_AMOUNT = 6

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename
        self.sheet = load_workbook(filename=filename, read_only=True).active

    def parse(self) -> Statement:
        """Main entry point for parsers

        super() implementation will call to split_records and parse_record to
        process the file.
        """

        account_id = self.sheet["A3"].value
        account_id = re.match(self.ACCOUNT_REGEX, account_id).group(1)
        assert account_id != None
        self.statement.account_id = account_id
        self.statement.currency = "SEK"
        self.statement.bank_id = "SJPRIO"

        logging.info(f"SELF f{self.statement}")
        rows = self.sheet.iter_rows()
        # Get headers
        headers = list(take(5, rows))
        assert [
            "Datum",
            "Bokfört",
            "Specifikation",
            "Ort",
            "Valuta",
            "Utl.belopp/moms",
            "Belopp",
        ] == [c.value for c in headers[-1]]
        # Remove last row with total
        self.rows = list(rows)[:-1]

        return super().parse()

    def split_records(self) -> Iterable[str]:
        """Return iterable object consisting of a line per transaction"""
        for row in self.rows:
            yield [c.value for c in row]

    def parse_record(self, line: str) -> StatementLine:
        """Parse given transaction line and return StatementLine object"""
        stmt_line = StatementLine()

        date = line[self.TRANSACTION_DATE]
        description = line[self.TRANSACTION_DESCRIPTION]
        amount = line[self.TRANSACTION_AMOUNT]
        location = line[self.TRANSACTION_LOCATION]

        stmt_line.date = date
        stmt_line.memo = f"{description} - {location}"
        stmt_line.amount = amount

        stmt_line.id = generate_transaction_id(stmt_line)

        return stmt_line
